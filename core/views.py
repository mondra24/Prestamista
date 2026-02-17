"""
Vistas del Sistema de Gestión de Préstamos
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.utils import timezone
from django.db.models import Sum, Count, Q
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from decimal import Decimal
import json

from .models import Cliente, Prestamo, Cuota, ConfiguracionMora, HistorialModificacionPago
from .forms import ClienteForm, PrestamoForm, RenovacionPrestamoForm


def fecha_local_hoy():
    """Retorna la fecha local (Argentina) en vez de UTC"""
    return timezone.localtime(timezone.now()).date()


def es_usuario_admin(user):
    """Verifica si el usuario es superusuario o tiene rol Administrador"""
    if user.is_superuser:
        return True
    return hasattr(user, 'perfil') and user.perfil.es_admin


def logout_view(request):
    """Vista para cerrar sesión"""
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('login')


class DashboardView(LoginRequiredMixin, TemplateView):
    """Vista principal del dashboard con resumen general"""
    template_name = 'core/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = fecha_local_hoy()
        
        # Filtro base por usuario (admin ve todo)
        cliente_filter = {}
        if not es_usuario_admin(self.request.user):
            cliente_filter['prestamo__cobrador'] = self.request.user
        
        # Estadísticas del día (incluye pagos completos y parciales)
        cobros_realizados_hoy = Cuota.objects.filter(
            fecha_pago_real=hoy,
            estado__in=['PA', 'PC'],
            **cliente_filter
        ).aggregate(
            total=Sum('monto_pagado'),
            cantidad=Count('id')
        )
        
        # Cuotas pendientes hoy
        cuotas_pendientes_hoy = Cuota.objects.filter(
            fecha_vencimiento=hoy,
            estado__in=['PE', 'PC'],
            prestamo__estado='AC',
            **cliente_filter
        ).count()
        
        # Cuotas vencidas total
        cuotas_vencidas = Cuota.objects.filter(
            fecha_vencimiento__lt=hoy,
            estado__in=['PE', 'PC'],
            prestamo__estado='AC',
            **cliente_filter
        ).count()
        
        # Total por cobrar hoy
        total_por_cobrar = Cuota.objects.filter(
            fecha_vencimiento=hoy,
            estado__in=['PE', 'PC'],
            prestamo__estado='AC',
            **cliente_filter
        ).aggregate(total=Sum('monto_cuota'))['total'] or Decimal('0.00')
        
        # Estadísticas generales (filtradas por usuario)
        if not es_usuario_admin(self.request.user):
            prestamos_activos = Prestamo.objects.filter(estado='AC', cobrador=self.request.user).count()
            clientes_activos = Cliente.objects.filter(estado='AC', usuario=self.request.user).count()
            total_cartera = Cuota.objects.filter(
                estado__in=['PE', 'PC'],
                prestamo__estado='AC',
                prestamo__cobrador=self.request.user
            ).aggregate(total=Sum('monto_cuota'))['total'] or Decimal('0.00')
        else:
            prestamos_activos = Prestamo.objects.filter(estado='AC').count()
            clientes_activos = Cliente.objects.filter(estado='AC').count()
            total_cartera = Cuota.objects.filter(
                estado__in=['PE', 'PC'],
                prestamo__estado='AC'
            ).aggregate(total=Sum('monto_cuota'))['total'] or Decimal('0.00')
        
        context.update({
            'total_cobrado_hoy': cobros_realizados_hoy['total'] or Decimal('0.00'),
            'cantidad_cobros_hoy': cobros_realizados_hoy['cantidad'] or 0,
            'cuotas_pendientes_hoy': cuotas_pendientes_hoy,
            'cuotas_vencidas': cuotas_vencidas,
            'total_por_cobrar': total_por_cobrar,
            'prestamos_activos': prestamos_activos,
            'clientes_activos': clientes_activos,
            'total_cartera': total_cartera,
            'fecha_hoy': hoy,
        })
        return context


class CobrosView(LoginRequiredMixin, TemplateView):
    """Vista de cobros del día con cuotas pendientes y vencidas"""
    template_name = 'core/cobros.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datetime import timedelta
        from .models import RutaCobro, ConfiguracionMora
        hoy = fecha_local_hoy()
        
        # Base queryset - filtrar por usuario si no es admin
        base_filter = {}
        if not es_usuario_admin(self.request.user):
            base_filter['prestamo__cobrador'] = self.request.user
        
        # Cuotas del día (pendientes) - ordenadas por ruta
        cuotas_hoy = Cuota.objects.filter(
            fecha_vencimiento=hoy,
            estado__in=['PE', 'PC'],
            prestamo__estado='AC',
            **base_filter
        ).select_related(
            'prestamo', 'prestamo__cliente', 'prestamo__cliente__ruta', 'prestamo__cliente__usuario', 'prestamo__cobrador'
        ).order_by(
            'prestamo__cliente__ruta__orden',
            'prestamo__cliente__ruta__nombre',
            'prestamo__cliente__apellido'
        )
        
        # Cuotas vencidas (días anteriores) - ordenadas por ruta y fecha
        cuotas_vencidas = Cuota.objects.filter(
            fecha_vencimiento__lt=hoy,
            estado__in=['PE', 'PC'],
            prestamo__estado='AC',
            **base_filter
        ).select_related(
            'prestamo', 'prestamo__cliente', 'prestamo__cliente__ruta', 'prestamo__cliente__usuario', 'prestamo__cobrador'
        ).order_by(
            'prestamo__cliente__ruta__orden',
            'prestamo__cliente__ruta__nombre',
            'fecha_vencimiento'
        )
        
        # Cuotas próximas (próximos 30 días) - ordenadas por fecha y ruta
        cuotas_proximas = Cuota.objects.filter(
            fecha_vencimiento__gt=hoy,
            fecha_vencimiento__lte=hoy + timedelta(days=30),
            estado__in=['PE', 'PC'],
            prestamo__estado='AC',
            **base_filter
        ).select_related(
            'prestamo', 'prestamo__cliente', 'prestamo__cliente__ruta', 'prestamo__cliente__usuario', 'prestamo__cobrador'
        ).order_by(
            'fecha_vencimiento',
            'prestamo__cliente__ruta__orden',
            'prestamo__cliente__ruta__nombre'
        )
        
        # Separar próxima semana de resto del mes
        cuotas_semana = [c for c in cuotas_proximas if c.fecha_vencimiento <= hoy + timedelta(days=7)]
        cuotas_mes = [c for c in cuotas_proximas if c.fecha_vencimiento > hoy + timedelta(days=7)]
        
        # Estadísticas del día
        cobros_filter = {'fecha_pago_real': hoy, 'estado__in': ['PA', 'PC']}
        if not es_usuario_admin(self.request.user):
            cobros_filter['prestamo__cobrador'] = self.request.user
        
        cobros_realizados_hoy = Cuota.objects.filter(
            **cobros_filter
        ).aggregate(
            total=Sum('monto_pagado'),
            cantidad=Count('id')
        )
        
        # Total por cobrar hoy
        total_por_cobrar = cuotas_hoy.aggregate(
            total=Sum('monto_cuota')
        )['total'] or Decimal('0.00')
        
        # Total próximos
        total_proximas = cuotas_proximas.aggregate(
            total=Sum('monto_cuota')
        )['total'] or Decimal('0.00')
        
        # Total vencidas
        total_vencidas = cuotas_vencidas.aggregate(
            total=Sum('monto_cuota')
        )['total'] or Decimal('0.00')
        
        # Obtener rutas activas para filtrado
        rutas = RutaCobro.objects.filter(activa=True).order_by('orden', 'nombre')
        
        # Obtener configuración de mora
        config_mora = ConfiguracionMora.obtener_config_activa()
        
        context.update({
            'cuotas_hoy': cuotas_hoy,
            'cuotas_vencidas': cuotas_vencidas,
            'cuotas_proximas': cuotas_proximas,
            'cuotas_semana': cuotas_semana,
            'cuotas_mes': cuotas_mes,
            'total_cobrado_hoy': cobros_realizados_hoy['total'] or Decimal('0.00'),
            'cantidad_cobros_hoy': cobros_realizados_hoy['cantidad'] or 0,
            'total_por_cobrar': total_por_cobrar,
            'total_proximas': total_proximas,
            'total_vencidas': total_vencidas,
            'fecha_hoy': hoy,
            'rutas': rutas,
            'config_mora': config_mora,
        })
        
        # Anotar historial de modificaciones en todas las cuotas
        from itertools import chain
        todas_cuotas = list(chain(cuotas_vencidas, cuotas_hoy, cuotas_semana, cuotas_mes))
        cuota_ids = [c.id for c in todas_cuotas]
        if cuota_ids:
            historiales = HistorialModificacionPago.objects.filter(
                cuota_id__in=cuota_ids
            ).select_related('cuota_relacionada', 'usuario').order_by('-fecha_modificacion')
            historial_map = {}
            for h in historiales:
                if h.cuota_id not in historial_map:
                    historial_map[h.cuota_id] = []
                historial_map[h.cuota_id].append(h)
            for cuota in todas_cuotas:
                cuota.historial_list = historial_map.get(cuota.id, [])
        
        return context


# ============== VISTAS DE CLIENTES ==============

class ClienteListView(LoginRequiredMixin, ListView):
    """Lista de clientes"""
    model = Cliente
    template_name = 'core/cliente_list.html'
    context_object_name = 'clientes'
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('usuario')
        
        # Filtrar por usuario (admin ve todos, otros solo los suyos)
        if not es_usuario_admin(self.request.user):
            queryset = queryset.filter(usuario=self.request.user)
        
        busqueda = self.request.GET.get('q', '')
        categoria = self.request.GET.get('categoria', '')
        
        if busqueda:
            queryset = queryset.filter(
                Q(nombre__icontains=busqueda) |
                Q(apellido__icontains=busqueda) |
                Q(telefono__icontains=busqueda)
            )
        
        if categoria:
            queryset = queryset.filter(categoria=categoria)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = Cliente.Categoria.choices
        return context


class ClienteCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo cliente"""
    model = Cliente
    form_class = ClienteForm
    template_name = 'core/cliente_form.html'
    success_url = reverse_lazy('core:cliente_list')
    
    def form_valid(self, form):
        # Asignar el usuario actual al cliente
        form.instance.usuario = self.request.user
        messages.success(self.request, 'Cliente creado exitosamente.')
        return super().form_valid(form)


class ClienteUpdateView(LoginRequiredMixin, UpdateView):
    """Editar cliente"""
    model = Cliente
    form_class = ClienteForm
    template_name = 'core/cliente_form.html'
    success_url = reverse_lazy('core:cliente_list')
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Admin puede editar todos, otros solo los suyos
        if not es_usuario_admin(self.request.user):
            queryset = queryset.filter(usuario=self.request.user)
        return queryset
    
    def form_valid(self, form):
        messages.success(self.request, 'Cliente actualizado exitosamente.')
        return super().form_valid(form)


class ClienteDetailView(LoginRequiredMixin, DetailView):
    """Detalle de cliente"""
    model = Cliente
    template_name = 'core/cliente_detail.html'
    context_object_name = 'cliente'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Admin puede ver todos, otros solo los suyos
        if not es_usuario_admin(self.request.user):
            queryset = queryset.filter(usuario=self.request.user)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['prestamos'] = self.object.prestamos.select_related('cobrador').all()
        context['prestamos_activos'] = self.object.prestamos.filter(
            estado='AC'
        ).select_related('cobrador')
        return context


# ============== VISTAS DE PRÉSTAMOS ==============

class PrestamoListView(LoginRequiredMixin, ListView):
    """Lista de préstamos"""
    model = Prestamo
    template_name = 'core/prestamo_list.html'
    context_object_name = 'prestamos'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por clientes del usuario (admin ve todos)
        if not es_usuario_admin(self.request.user):
            queryset = queryset.filter(cobrador=self.request.user)
        
        estado = self.request.GET.get('estado', '')
        
        if estado:
            queryset = queryset.filter(estado=estado)
        
        return queryset.select_related('cliente', 'cliente__usuario', 'cobrador')


class PrestamoCreateView(LoginRequiredMixin, CreateView):
    """Crear nuevo préstamo"""
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'core/prestamo_form.html'
    success_url = reverse_lazy('core:prestamo_list')
    
    def get_initial(self):
        initial = super().get_initial()
        cliente_id = self.request.GET.get('cliente')
        if cliente_id:
            try:
                initial['cliente'] = int(cliente_id)
            except (ValueError, TypeError):
                pass
        initial['fecha_inicio'] = fecha_local_hoy()
        return initial
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrar clientes por usuario (admin ve todos)
        if not es_usuario_admin(self.request.user):
            form.fields['cliente'].queryset = Cliente.objects.filter(
                estado='AC', usuario=self.request.user
            ).order_by('apellido', 'nombre')
        else:
            form.fields['cliente'].queryset = Cliente.objects.filter(
                estado='AC'
            ).order_by('apellido', 'nombre')
        return form
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasar datos de clientes para mostrar límite de crédito
        if not es_usuario_admin(self.request.user):
            context['clientes'] = Cliente.objects.filter(
                estado='AC', usuario=self.request.user
            ).order_by('apellido', 'nombre')
        else:
            context['clientes'] = Cliente.objects.filter(
                estado='AC'
            ).order_by('apellido', 'nombre')
        return context
    
    def form_valid(self, form):
        # Asignar el cobrador actual al préstamo
        form.instance.cobrador = self.request.user
        messages.success(self.request, 'Préstamo creado exitosamente. Las cuotas han sido generadas.')
        return super().form_valid(form)


class PrestamoDetailView(LoginRequiredMixin, DetailView):
    """Detalle de préstamo con todas sus cuotas"""
    model = Prestamo
    template_name = 'core/prestamo_detail.html'
    context_object_name = 'prestamo'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Admin puede ver todos, otros solo los de sus préstamos
        if not es_usuario_admin(self.request.user):
            queryset = queryset.filter(cobrador=self.request.user)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cuotas = list(self.object.cuotas.all())
        context['config_mora'] = ConfiguracionMora.obtener_config_activa()
        
        # Cargar historial de modificaciones para todas las cuotas
        cuota_ids = [c.id for c in cuotas]
        historial = HistorialModificacionPago.objects.filter(
            cuota_id__in=cuota_ids
        ).select_related('cuota_relacionada', 'usuario').order_by('-fecha_modificacion')
        
        historial_por_cuota = {}
        for h in historial:
            if h.cuota_id not in historial_por_cuota:
                historial_por_cuota[h.cuota_id] = []
            historial_por_cuota[h.cuota_id].append(h)
        
        # Anotar cada cuota con su historial
        for cuota in cuotas:
            cuota.historial_list = historial_por_cuota.get(cuota.id, [])
        
        context['cuotas'] = cuotas
        return context


class RenovarPrestamoView(LoginRequiredMixin, TemplateView):
    """Vista para renovar un préstamo"""
    template_name = 'core/prestamo_renovar.html'
    
    def get_prestamo(self):
        # Verificar propiedad del préstamo
        if not es_usuario_admin(self.request.user):
            return get_object_or_404(Prestamo, pk=self.kwargs['pk'], cobrador=self.request.user)
        return get_object_or_404(Prestamo, pk=self.kwargs['pk'])
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        prestamo = self.get_prestamo()
        saldo_pendiente = prestamo.calcular_saldo_para_renovacion()
        context['prestamo'] = prestamo
        context['saldo_pendiente'] = saldo_pendiente
        context['form'] = kwargs.get('form', RenovacionPrestamoForm(
            cliente=prestamo.cliente,
            saldo_pendiente=saldo_pendiente,
            initial={
                'nueva_tasa': prestamo.tasa_interes_porcentaje,
                'nuevas_cuotas': prestamo.cuotas_pactadas,
                'nueva_frecuencia': prestamo.frecuencia,
            }
        ))
        # Agregar información del límite de crédito
        context['maximo_capital_adicional'] = prestamo.cliente.maximo_prestable
        if context['maximo_capital_adicional'] is not None:
            context['maximo_capital_adicional'] += saldo_pendiente
        return context
    
    def get(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())
    
    def post(self, request, *args, **kwargs):
        prestamo_anterior = self.get_prestamo()
        saldo_pendiente = prestamo_anterior.calcular_saldo_para_renovacion()
        form = RenovacionPrestamoForm(
            request.POST,
            cliente=prestamo_anterior.cliente,
            saldo_pendiente=saldo_pendiente
        )
        if form.is_valid():
            nuevo_prestamo = Prestamo.renovar_prestamo(
                prestamo_anterior=prestamo_anterior,
                nuevo_monto=form.cleaned_data['nuevo_monto'],
                nueva_tasa=form.cleaned_data['nueva_tasa'],
                nuevas_cuotas=form.cleaned_data['nuevas_cuotas'],
                nueva_frecuencia=form.cleaned_data['nueva_frecuencia'],
                cobrador=request.user
            )
            
            messages.success(
                request, 
                f'Préstamo renovado exitosamente. Nuevo préstamo #{nuevo_prestamo.pk} creado.'
            )
            return redirect('core:prestamo_detail', pk=nuevo_prestamo.pk)
        
        return self.render_to_response(self.get_context_data(form=form))


# ============== VISTAS DE COBROS (AJAX) ==============

@login_required
def cobrar_cuota(request, pk):
    """Registrar pago de cuota via AJAX"""
    if request.method == 'POST':
        try:
            # Solo el cobrador asignado puede cobrar
            cuota = get_object_or_404(Cuota, pk=pk, prestamo__cobrador=request.user)
            
            # Obtener datos del body
            try:
                data = json.loads(request.body)
                monto = Decimal(str(data.get('monto', cuota.monto_restante)))
                accion_restante = data.get('accion_restante', 'ignorar')  # 'ignorar', 'proxima', 'especial'
                fecha_especial_str = data.get('fecha_especial', None)
                
                # Nuevos campos de método de pago
                metodo_pago = data.get('metodo_pago', 'EF')  # 'EF', 'TR', 'MX'
                monto_efectivo = data.get('monto_efectivo')
                monto_transferencia = data.get('monto_transferencia')
                referencia_transferencia = data.get('referencia_transferencia')
                
                # Interés por mora
                interes_mora = data.get('interes_mora', 0)
                
                # Convertir fecha especial si existe
                fecha_especial = None
                if fecha_especial_str and accion_restante == 'especial':
                    from datetime import datetime
                    fecha_especial = datetime.strptime(fecha_especial_str, '%Y-%m-%d').date()
                    
            except (json.JSONDecodeError, ValueError):
                monto = None
                accion_restante = 'ignorar'
                fecha_especial = None
                metodo_pago = 'EF'
                monto_efectivo = None
                monto_transferencia = None
                referencia_transferencia = None
                interes_mora = 0
            
            # Calcular restante antes del pago
            monto_restante_antes = float(cuota.monto_restante)
            monto_que_quedara = max(0, monto_restante_antes - float(monto or cuota.monto_restante))
            
            cuota.registrar_pago(
                monto=monto, 
                accion_restante=accion_restante, 
                fecha_especial=fecha_especial,
                metodo_pago=metodo_pago,
                monto_efectivo=monto_efectivo,
                monto_transferencia=monto_transferencia,
                referencia_transferencia=referencia_transferencia,
                interes_mora=interes_mora,
                cobrador=request.user
            )
            
            # Mensaje según la acción
            if accion_restante == 'proxima' and monto_que_quedara > 0:
                mensaje = f'Pago registrado. ${monto_que_quedara:.2f} sumado a la próxima cuota.'
            elif accion_restante == 'especial' and monto_que_quedara > 0:
                mensaje = f'Pago registrado. Cuota especial creada por ${monto_que_quedara:.2f}.'
            else:
                mensaje = 'Pago registrado exitosamente'
            
            # Agregar info de método de pago al mensaje
            if metodo_pago == 'TR':
                mensaje += ' (Transferencia)'
            elif metodo_pago == 'MX':
                mensaje += ' (Mixto)'
            
            # Calcular total cobrado hoy (incluye pagos parciales)
            hoy = fecha_local_hoy()
            stats_filter = {
                'fecha_pago_real': hoy,
                'estado__in': ['PA', 'PC'],
            }
            if not es_usuario_admin(request.user):
                stats_filter['prestamo__cobrador'] = request.user
            
            total_cobrado_hoy = Cuota.objects.filter(
                **stats_filter
            ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
            
            cantidad_cobros_hoy = Cuota.objects.filter(
                **stats_filter
            ).count()
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cuota': {
                    'id': cuota.pk,
                    'estado': cuota.estado,
                    'estado_display': cuota.get_estado_display(),
                    'monto_pagado': float(cuota.monto_pagado),
                    'monto_restante': float(cuota.monto_restante),
                    'metodo_pago': cuota.metodo_pago,
                    'interes_mora_cobrado': float(cuota.interes_mora_cobrado),
                },
                'prestamo': {
                    'progreso': cuota.prestamo.progreso_porcentaje,
                    'estado': cuota.prestamo.estado,
                },
                'estadisticas': {
                    'total_cobrado_hoy': int(total_cobrado_hoy),  # Sin decimales
                    'cantidad_cobros_hoy': cantidad_cobros_hoy,
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required
def obtener_cuotas_hoy(request):
    """Obtener cuotas del día via AJAX (para actualización en tiempo real)"""
    hoy = fecha_local_hoy()
    
    cuotas_qs = Cuota.objects.filter(
        fecha_vencimiento=hoy,
        estado__in=['PE', 'PC'],
        prestamo__estado='AC'
    )
    # Filtrar por usuario (admin ve todo)
    if not es_usuario_admin(request.user):
        cuotas_qs = cuotas_qs.filter(prestamo__cobrador=request.user)
    
    cuotas = cuotas_qs.select_related('prestamo', 'prestamo__cliente', 'prestamo__cliente__usuario', 'prestamo__cobrador').values(
        'id', 'numero_cuota', 'monto_cuota', 'estado',
        'prestamo__id', 'prestamo__cuotas_pactadas',
        'prestamo__cliente__nombre', 'prestamo__cliente__apellido',
        'prestamo__cobrador__username', 'prestamo__cobrador__first_name',
        'prestamo__cobrador__last_name'
    )
    
    return JsonResponse({
        'cuotas': list(cuotas)
    })


@login_required
def cambiar_categoria_cliente(request, pk):
    """Cambiar categoría del cliente via AJAX"""
    if request.method == 'POST':
        try:
            # Verificar propiedad del cliente
            if not es_usuario_admin(request.user):
                cliente = get_object_or_404(Cliente, pk=pk, usuario=request.user)
            else:
                cliente = get_object_or_404(Cliente, pk=pk)
            
            data = json.loads(request.body)
            nueva_categoria = data.get('categoria')
            
            if nueva_categoria not in ['EX', 'RE', 'MO', 'NU']:
                return JsonResponse({
                    'success': False,
                    'message': 'Categoría no válida'
                }, status=400)
            
            categoria_anterior = cliente.get_categoria_display()
            cliente.categoria = nueva_categoria
            cliente.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Categoría cambiada de {categoria_anterior} a {cliente.get_categoria_display()}',
                'cliente': {
                    'id': cliente.pk,
                    'categoria': cliente.categoria,
                    'categoria_display': cliente.get_categoria_display(),
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required
def buscar_clientes(request):
    """Búsqueda de clientes via AJAX para autocompletado"""
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})
    
    queryset = Cliente.objects.filter(estado='AC')
    if not es_usuario_admin(request.user):
        queryset = queryset.filter(usuario=request.user)
    
    queryset = queryset.filter(
        Q(nombre__icontains=q) |
        Q(apellido__icontains=q) |
        Q(telefono__icontains=q)
    ).select_related('ruta')[:15]
    
    results = []
    for c in queryset:
        results.append({
            'id': c.pk,
            'nombre': c.nombre_completo,
            'telefono': c.telefono or '',
            'ruta': c.ruta.nombre if c.ruta else '',
            'categoria': c.categoria,
            'categoria_display': c.get_categoria_display(),
        })
    
    return JsonResponse({'results': results})


# ============== VISTAS DE REPORTES ==============

class CierreCajaView(LoginRequiredMixin, TemplateView):
    """Vista de cierre de caja del día"""
    template_name = 'core/cierre_caja.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener fecha del filtro o usar hoy
        fecha_str = self.request.GET.get('fecha')
        if fecha_str:
            from datetime import datetime
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            fecha = fecha_local_hoy()
        
        # Pagos del día (incluye pagos completos y parciales)
        pagos_del_dia = Cuota.objects.filter(
            fecha_pago_real=fecha,
            estado__in=['PA', 'PC']
        )
        # Filtrar por usuario (admin ve todo)
        if not es_usuario_admin(self.request.user):
            pagos_del_dia = pagos_del_dia.filter(prestamo__cobrador=self.request.user)
        pagos_del_dia = pagos_del_dia.select_related('prestamo', 'prestamo__cliente', 'cobrado_por').order_by(
            'prestamo__cliente__apellido'
        )
        
        # Total cobrado
        total_cobrado = pagos_del_dia.aggregate(
            total=Sum('monto_pagado')
        )['total'] or Decimal('0.00')
        
        context.update({
            'fecha': fecha,
            'pagos': pagos_del_dia,
            'total_cobrado': total_cobrado,
            'cantidad_pagos': pagos_del_dia.count(),
        })
        
        # Anotar historial de modificaciones en cada pago
        pagos_list = list(pagos_del_dia)
        pago_ids = [p.id for p in pagos_list]
        if pago_ids:
            historiales = HistorialModificacionPago.objects.filter(
                cuota_id__in=pago_ids
            ).select_related('cuota_relacionada', 'usuario').order_by('-fecha_modificacion')
            historial_map = {}
            for h in historiales:
                if h.cuota_id not in historial_map:
                    historial_map[h.cuota_id] = []
                historial_map[h.cuota_id].append(h)
            for pago in pagos_list:
                pago.historial_list = historial_map.get(pago.id, [])
        context['pagos'] = pagos_list
        
        return context


class PlanillaImpresionView(LoginRequiredMixin, TemplateView):
    """Vista optimizada para impresión con cuotas pendientes del día"""
    template_name = 'core/planilla_impresion.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from collections import OrderedDict
        from .models import RutaCobro, ConfiguracionPlanilla, ColumnaPlanilla
        
        # Obtener configuración de planilla
        config_id = self.request.GET.get('config')
        if config_id:
            try:
                config = ConfiguracionPlanilla.objects.get(pk=config_id)
            except ConfiguracionPlanilla.DoesNotExist:
                config = ConfiguracionPlanilla.objects.filter(es_default=True).first()
        else:
            config = ConfiguracionPlanilla.objects.filter(es_default=True).first()
        
        # Si no hay configuración, usar valores por defecto
        if not config:
            config = type('ConfigDefault', (), {
                'titulo_reporte': 'PLANILLA DE COBROS',
                'subtitulo': 'Préstamos - Sistema de Gestión',
                'mostrar_logo': False,
                'mostrar_fecha': True,
                'mostrar_totales': True,
                'mostrar_firmas': True,
                'agrupar_por_ruta': True,
                'agrupar_por_categoria': False,
                'incluir_vencidas': True,  # Incluir vencidas por defecto
                'filtrar_por_ruta': None,
                'pk': None,
            })()
        
        # Obtener columnas activas
        columnas = ColumnaPlanilla.objects.filter(activa=True).order_by('orden')
        if not columnas.exists():
            # Crear columnas por defecto si no existen
            columnas_default = [
                {'nombre_columna': 'numero', 'titulo_personalizado': '#', 'orden': 1, 'ancho': '4%'},
                {'nombre_columna': 'nombre_cliente', 'titulo_personalizado': 'Cliente', 'orden': 2, 'ancho': '22%'},
                {'nombre_columna': 'telefono', 'titulo_personalizado': 'Teléfono', 'orden': 3, 'ancho': '12%'},
                {'nombre_columna': 'categoria', 'titulo_personalizado': 'Cat.', 'orden': 4, 'ancho': '8%'},
                {'nombre_columna': 'cuota_actual', 'titulo_personalizado': 'Cuota', 'orden': 5, 'ancho': '10%'},
                {'nombre_columna': 'monto_cuota', 'titulo_personalizado': 'Monto', 'orden': 6, 'ancho': '12%'},
                {'nombre_columna': 'es_renovacion', 'titulo_personalizado': 'Renov.', 'orden': 7, 'ancho': '10%'},
                {'nombre_columna': 'dia_pago', 'titulo_personalizado': 'Día Pago', 'orden': 8, 'ancho': '10%'},
                {'nombre_columna': 'espacio_cobrado', 'titulo_personalizado': 'Cobrado', 'orden': 9, 'ancho': '12%'},
            ]
            for col_data in columnas_default:
                ColumnaPlanilla.objects.create(**col_data)
            columnas = ColumnaPlanilla.objects.filter(activa=True).order_by('orden')
        
        fecha_str = self.request.GET.get('fecha')
        if fecha_str:
            from datetime import datetime
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            fecha = fecha_local_hoy()
        
        ruta_id = self.request.GET.get('ruta')
        if not ruta_id and config and hasattr(config, 'filtrar_por_ruta') and config.filtrar_por_ruta:
            ruta_id = config.filtrar_por_ruta_id
        
        # Verificar si incluir vencidas (por defecto True para mostrar todos los pendientes)
        incluir_vencidas_param = self.request.GET.get('incluir_vencidas')
        if incluir_vencidas_param is not None:
            incluir_vencidas = incluir_vencidas_param.lower() in ('true', '1', 'si', 'yes')
        else:
            incluir_vencidas = getattr(config, 'incluir_vencidas', True)
        
        # Verificar si mostrar próximas cuotas
        mostrar_proximas = self.request.GET.get('proximas', 'true').lower() in ('true', '1', 'si', 'yes')
        
        # Verificar si es modo cierre de caja (mostrar cobros realizados)
        tipo_planilla = self.request.GET.get('tipo', '')
        es_cierre = tipo_planilla == 'cierre'
        
        if es_cierre:
            # MODO CIERRE: Mostrar cuotas COBRADAS en la fecha (completas y parciales)
            cuotas_pendientes = Cuota.objects.filter(
                fecha_pago_real=fecha,
                estado__in=['PA', 'PC']
            )
            if not es_usuario_admin(self.request.user):
                cuotas_pendientes = cuotas_pendientes.filter(prestamo__cobrador=self.request.user)
            cuotas_pendientes = cuotas_pendientes.select_related(
                'prestamo',
                'prestamo__cliente',
                'prestamo__cliente__ruta',
                'prestamo__cliente__tipo_negocio'
            ).order_by('prestamo__cliente__apellido')
            
            # Cambiar título para cierre
            config.titulo_reporte = 'CIERRE DE CAJA'
            config.subtitulo = f'Cobros realizados el {fecha.strftime("%d/%m/%Y")}'
        else:
            # MODO NORMAL: Obtener cuotas PENDIENTES (no cobradas)
            estados_cuota = ['PE', 'PC']
        
            # Base query: cuotas pendientes de préstamos activos
            base_query = Cuota.objects.filter(
                estado__in=estados_cuota,
                prestamo__estado='AC'
            )
            if not es_usuario_admin(self.request.user):
                base_query = base_query.filter(prestamo__cobrador=self.request.user)
            base_query = base_query.select_related(
                'prestamo', 
                'prestamo__cliente', 
                'prestamo__cliente__ruta',
                'prestamo__cliente__tipo_negocio'
            )
            
            if incluir_vencidas and mostrar_proximas:
                # Mostrar: cuotas vencidas + cuotas del día seleccionado + próximas 7 días
                from datetime import timedelta
                fecha_limite = fecha + timedelta(days=7)
                cuotas_pendientes = base_query.filter(
                    fecha_vencimiento__lte=fecha_limite
                )
            elif incluir_vencidas:
                # Mostrar: cuotas vencidas + cuotas del día seleccionado
                cuotas_pendientes = base_query.filter(
                    fecha_vencimiento__lte=fecha
                )
            elif mostrar_proximas:
                # Mostrar: solo cuotas de la fecha seleccionada + próximos 7 días
                from datetime import timedelta
                fecha_limite = fecha + timedelta(days=7)
                cuotas_pendientes = base_query.filter(
                    fecha_vencimiento__gte=fecha,
                    fecha_vencimiento__lte=fecha_limite
                )
            else:
                # Solo cuotas del día exacto
                cuotas_pendientes = base_query.filter(
                    fecha_vencimiento=fecha
                )
        
        # Ordenar
        if hasattr(config, 'agrupar_por_ruta') and config.agrupar_por_ruta:
            cuotas_pendientes = cuotas_pendientes.order_by(
                'prestamo__cliente__ruta__orden',
                'prestamo__cliente__apellido'
            )
        elif hasattr(config, 'agrupar_por_categoria') and config.agrupar_por_categoria:
            cuotas_pendientes = cuotas_pendientes.order_by(
                'prestamo__cliente__categoria',
                'prestamo__cliente__apellido'
            )
        else:
            cuotas_pendientes = cuotas_pendientes.order_by('prestamo__cliente__apellido')
        
        # Filtrar por ruta si se especifica
        ruta_filter = None
        if ruta_id:
            try:
                ruta_filter = RutaCobro.objects.get(pk=ruta_id)
                cuotas_pendientes = cuotas_pendientes.filter(
                    prestamo__cliente__ruta_id=ruta_id
                )
            except RutaCobro.DoesNotExist:
                pass
        
        # Organizar datos
        cuotas_agrupadas = OrderedDict()
        if hasattr(config, 'agrupar_por_ruta') and config.agrupar_por_ruta:
            for cuota in cuotas_pendientes:
                grupo = cuota.prestamo.cliente.ruta.nombre if cuota.prestamo.cliente.ruta else "Sin Ruta"
                if grupo not in cuotas_agrupadas:
                    cuotas_agrupadas[grupo] = []
                cuotas_agrupadas[grupo].append(cuota)
        elif hasattr(config, 'agrupar_por_categoria') and config.agrupar_por_categoria:
            for cuota in cuotas_pendientes:
                grupo = cuota.prestamo.cliente.get_categoria_display()
                if grupo not in cuotas_agrupadas:
                    cuotas_agrupadas[grupo] = []
                cuotas_agrupadas[grupo].append(cuota)
        else:
            cuotas_agrupadas['Todos'] = list(cuotas_pendientes)
        
        total_esperado = cuotas_pendientes.aggregate(
            total=Sum('monto_pagado' if es_cierre else 'monto_cuota')
        )['total'] or Decimal('0.00')
        
        # Lista de rutas para filtro
        rutas = RutaCobro.objects.filter(activa=True).order_by('orden')
        
        # Lista de configuraciones disponibles
        configuraciones = ConfiguracionPlanilla.objects.all()
        
        context.update({
            'fecha': fecha,
            'cuotas_pendientes': cuotas_pendientes,
            'cuotas_por_ruta': cuotas_agrupadas,  # Compatibilidad
            'cuotas_agrupadas': cuotas_agrupadas,
            'total_esperado': total_esperado,
            'rutas': rutas,
            'ruta_filter': ruta_filter,
            'now': timezone.now(),
            'config': config,
            'columnas': columnas,
            'configuraciones': configuraciones,
            'incluir_vencidas': incluir_vencidas,
            'mostrar_proximas': mostrar_proximas,
            'es_cierre': es_cierre,
        })
        return context


class ReporteGeneralView(LoginRequiredMixin, TemplateView):
    """Vista con reportes generales"""
    template_name = 'core/reporte_general.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas generales - filtradas por usuario
        if not es_usuario_admin(self.request.user):
            clientes_qs = Cliente.objects.filter(estado='AC', usuario=self.request.user)
            prestamos_qs = Prestamo.objects.filter(estado='AC', cobrador=self.request.user)
            cuotas_qs = Cuota.objects.filter(prestamo__estado='AC', prestamo__cobrador=self.request.user)
        else:
            clientes_qs = Cliente.objects.filter(estado='AC')
            prestamos_qs = Prestamo.objects.filter(estado='AC')
            cuotas_qs = Cuota.objects.filter(prestamo__estado='AC')
        
        context['total_clientes'] = clientes_qs.count()
        context['prestamos_activos'] = prestamos_qs.count()
        
        # Capital en la calle (monto pendiente de todos los préstamos activos)
        capital_calle = sum(p.monto_pendiente for p in prestamos_qs)
        context['capital_en_calle'] = capital_calle
        
        # Cuotas vencidas
        hoy = fecha_local_hoy()
        context['cuotas_vencidas'] = cuotas_qs.filter(
            fecha_vencimiento__lt=hoy,
            estado__in=['PE', 'PC'],
        ).count()
        
        # Distribución por categoría de clientes
        context['clientes_por_categoria'] = clientes_qs.values('categoria').annotate(
            cantidad=Count('id')
        )
        
        return context


# ============== VISTAS DE GESTIÓN DE USUARIOS ==============

from django.contrib.auth.models import User
from .models import PerfilUsuario
from .forms import UsuarioForm, UsuarioEditForm


class UsuarioListView(LoginRequiredMixin, ListView):
    """Lista de usuarios del sistema"""
    model = User
    template_name = 'core/usuario_list.html'
    context_object_name = 'usuarios'
    
    def dispatch(self, request, *args, **kwargs):
        # Solo admins pueden ver usuarios
        if not es_usuario_admin(request.user):
            messages.error(request, 'No tienes permiso para acceder a esta sección.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return User.objects.select_related('perfil').order_by('-date_joined')


class UsuarioCreateView(LoginRequiredMixin, TemplateView):
    """Crear nuevo usuario"""
    template_name = 'core/usuario_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        # Solo admins pueden crear usuarios
        if not es_usuario_admin(request.user):
            messages.error(request, 'No tienes permiso para crear usuarios.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = kwargs.get('form', UsuarioForm())
        context['titulo'] = 'Nuevo Usuario'
        return context
    
    def get(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())
    
    def post(self, request, *args, **kwargs):
        form = UsuarioForm(request.POST)
        if form.is_valid():
            # Crear usuario
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
            )
            
            # Actualizar perfil
            if hasattr(user, 'perfil'):
                user.perfil.rol = form.cleaned_data['rol']
                user.perfil.telefono = form.cleaned_data['telefono']
                user.perfil.save()
            else:
                PerfilUsuario.objects.create(
                    user=user,
                    rol=form.cleaned_data['rol'],
                    telefono=form.cleaned_data['telefono']
                )
            
            messages.success(request, f'Usuario "{user.username}" creado exitosamente.')
            return redirect('core:usuario_list')
        
        return self.render_to_response(self.get_context_data(form=form))


class UsuarioEditView(LoginRequiredMixin, TemplateView):
    """Editar usuario existente"""
    template_name = 'core/usuario_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        # Solo admins pueden editar usuarios
        if not es_usuario_admin(request.user):
            messages.error(request, 'No tienes permiso para editar usuarios.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_user(self):
        return get_object_or_404(User, pk=self.kwargs['pk'])
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.get_user()
        context['usuario_edit'] = user
        context['titulo'] = f'Editar Usuario: {user.username}'
        
        if 'form' not in kwargs:
            initial = {
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'rol': user.perfil.rol if hasattr(user, 'perfil') else 'CO',
                'telefono': user.perfil.telefono if hasattr(user, 'perfil') else '',
                'activo': user.perfil.activo if hasattr(user, 'perfil') else True,
            }
            context['form'] = UsuarioEditForm(initial=initial)
        else:
            context['form'] = kwargs['form']
        
        return context
    
    def get(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())
    
    def post(self, request, *args, **kwargs):
        form = UsuarioEditForm(request.POST)
        user = self.get_user()
        
        if form.is_valid():
            # Actualizar usuario
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data['email']
            
            # Cambiar contraseña si se proporcionó
            if form.cleaned_data['password']:
                user.set_password(form.cleaned_data['password'])
            
            user.save()
            
            # Actualizar perfil
            if hasattr(user, 'perfil'):
                user.perfil.rol = form.cleaned_data['rol']
                user.perfil.telefono = form.cleaned_data['telefono']
                user.perfil.activo = form.cleaned_data['activo']
                user.perfil.save()
            
            messages.success(request, f'Usuario "{user.username}" actualizado exitosamente.')
            return redirect('core:usuario_list')
        
        return self.render_to_response(self.get_context_data(form=form))


@login_required
def toggle_usuario_activo(request, pk):
    """Activar/desactivar un usuario"""
    # Verificar permisos
    if not es_usuario_admin(request.user):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('core:dashboard')
    
    user = get_object_or_404(User, pk=pk)
    
    # No permitir desactivarse a sí mismo
    if user == request.user:
        messages.error(request, 'No puedes desactivarte a ti mismo.')
        return redirect('core:usuario_list')
    
    if hasattr(user, 'perfil'):
        user.perfil.activo = not user.perfil.activo
        user.perfil.save()
        user.is_active = user.perfil.activo
        user.save()
        
        estado = 'activado' if user.perfil.activo else 'desactivado'
        messages.success(request, f'Usuario "{user.username}" {estado}.')
    
    return redirect('core:usuario_list')


# ==================== EXPORTACIÓN EXCEL ====================

from django.http import HttpResponse
from .models import RegistroAuditoria, Notificacion, ConfiguracionRespaldo
import io
import os
from datetime import datetime


@login_required
def exportar_planilla_excel(request):
    """Exportar planilla de cobros a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La exportación a Excel no está disponible. Instale openpyxl.')
        return redirect('core:planilla_impresion')
    
    # Obtener fecha del filtro
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha = fecha_local_hoy()
    
    # Obtener ruta de filtro
    ruta_id = request.GET.get('ruta')
    incluir_vencidas = request.GET.get('incluir_vencidas', '1').lower() in ('true', '1', 'si')
    
    # Obtener cuotas pendientes
    cuotas = Cuota.objects.filter(
        prestamo__estado='AC',
        estado__in=['PE', 'PC']
    )
    if not es_usuario_admin(request.user):
        cuotas = cuotas.filter(prestamo__cobrador=request.user)
    cuotas = cuotas.select_related('prestamo', 'prestamo__cliente', 'prestamo__cliente__ruta')
    
    if incluir_vencidas:
        cuotas = cuotas.filter(fecha_vencimiento__lte=fecha)
    else:
        cuotas = cuotas.filter(fecha_vencimiento=fecha)
    
    if ruta_id:
        cuotas = cuotas.filter(prestamo__cliente__ruta_id=ruta_id)
    
    cuotas = cuotas.order_by('prestamo__cliente__ruta__orden', 'prestamo__cliente__apellido')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Planilla {fecha.strftime('%d-%m-%Y')}"
    
    # Pre-cargar historial de modificaciones (cuotas que recibieron monto)
    cuota_ids = list(cuotas.values_list('id', flat=True))
    cuotas_con_monto_recibido = {}
    if cuota_ids:
        historiales_recibidos = HistorialModificacionPago.objects.filter(
            cuota_id__in=cuota_ids,
            tipo_modificacion='MR'
        ).select_related('cuota_relacionada')
        for h in historiales_recibidos:
            cuotas_con_monto_recibido[h.cuota_id] = h
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    recibida_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Celeste claro
    
    # Título
    ws.merge_cells('A1:M1')
    ws['A1'] = f'PLANILLA DE COBROS - {fecha.strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Info
    ws.merge_cells('A2:M2')
    ws['A2'] = f'Total cobros: {cuotas.count()} | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Leyenda
    ws.merge_cells('A3:M3')
    ws['A3'] = '■ Celeste = Cuota modificada (recibió monto de otra cuota por pago parcial)'
    ws['A3'].font = Font(italic=True, size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['#', 'Préstamo', 'Cliente', 'Teléfono', 'Ruta', 'Cuota', 'Monto', 'Monto Original', 'Venc.', 'Fecha Fin', 'Cobrado', 'Modificada', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 40
    
    # Datos
    total = Decimal('0.00')
    for i, cuota in enumerate(cuotas, 1):
        row = i + 5
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=f'#{cuota.prestamo.pk}').border = border
        ws.cell(row=row, column=3, value=cuota.prestamo.cliente.nombre_completo).border = border
        ws.cell(row=row, column=4, value=cuota.prestamo.cliente.telefono).border = border
        ws.cell(row=row, column=5, value=cuota.prestamo.cliente.ruta.nombre if cuota.prestamo.cliente.ruta else 'Sin Ruta').border = border
        ws.cell(row=row, column=6, value=f'{cuota.numero_cuota}/{cuota.prestamo.cuotas_pactadas}').border = border
        
        monto_cell = ws.cell(row=row, column=7, value=float(cuota.monto_cuota))
        monto_cell.number_format = '#,##0'
        monto_cell.border = border
        
        # Columna Monto Original y Observaciones
        recibido = cuotas_con_monto_recibido.get(cuota.id)
        fue_modificada = recibido is not None
        
        if recibido:
            orig_cell = ws.cell(row=row, column=8, value=float(recibido.monto_cuota_anterior))
            orig_cell.number_format = '#,##0'
        else:
            orig_cell = ws.cell(row=row, column=8, value='-')
        orig_cell.border = border
        
        ws.cell(row=row, column=9, value=cuota.fecha_vencimiento.strftime('%d/%m')).border = border
        ws.cell(row=row, column=10, value=cuota.prestamo.fecha_finalizacion.strftime('%d/%m/%Y') if cuota.prestamo.fecha_finalizacion else '-').border = border
        ws.cell(row=row, column=11, value='').border = border
        
        mod_cell = ws.cell(row=row, column=12, value='SÍ' if fue_modificada else '-')
        mod_cell.border = border
        mod_cell.alignment = Alignment(horizontal='center')
        if fue_modificada:
            mod_cell.font = Font(bold=True, color='856404')
        
        obs_text = ''
        if recibido:
            origen = f' de cuota #{recibido.cuota_relacionada.numero_cuota}' if recibido.cuota_relacionada else ''
            obs_text = f'Recibió ${recibido.monto_restante_transferido:,.0f}{origen}'
            if recibido.interes_mora > 0:
                obs_text += f' (mora: ${recibido.interes_mora:,.0f})'
        
        obs_cell = ws.cell(row=row, column=13, value=obs_text if obs_text else '-')
        obs_cell.border = border
        obs_cell.alignment = Alignment(wrap_text=True)
        
        # Aplicar color de fondo si fue modificada
        if fue_modificada:
            for col_idx in range(1, 14):
                ws.cell(row=row, column=col_idx).fill = recibida_fill
        
        total += cuota.monto_cuota
    
    # Fila de total
    total_row = cuotas.count() + 6
    ws.merge_cells(f'A{total_row}:G{total_row}')
    ws.cell(row=total_row, column=1, value='TOTAL ESPERADO:').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=8, value=float(total))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0'
    
    # Registrar auditoría
    RegistroAuditoria.registrar(
        usuario=request.user,
        tipo_accion='OT',
        tipo_modelo='SI',
        descripcion=f'Exportación de planilla a Excel - Fecha: {fecha}',
        ip_address=get_client_ip(request)
    )
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=planilla_cobros_{fecha.strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def exportar_cierre_excel(request):
    """Exportar cierre de caja a Excel con cobros realizados"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La exportación a Excel no está disponible. Instale openpyxl.')
        return redirect('core:cierre_caja')
    
    # Obtener fecha
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha = fecha_local_hoy()
    
    # Obtener cobros del día (completos y parciales)
    pagos = Cuota.objects.filter(
        fecha_pago_real=fecha,
        estado__in=['PA', 'PC']
    )
    if not es_usuario_admin(request.user):
        pagos = pagos.filter(prestamo__cobrador=request.user)
    pagos = pagos.select_related('prestamo', 'prestamo__cliente', 'prestamo__cliente__ruta', 'cobrado_por').order_by(
        'prestamo__cliente__apellido'
    )
    
    # Pre-cargar historial de modificaciones para las cuotas del día
    cuota_ids = list(pagos.values_list('id', flat=True))
    historial_por_cuota = {}
    if cuota_ids:
        historiales = HistorialModificacionPago.objects.filter(
            cuota_id__in=cuota_ids
        ).select_related('cuota_relacionada').order_by('fecha_modificacion')
        for h in historiales:
            if h.cuota_id not in historial_por_cuota:
                historial_por_cuota[h.cuota_id] = []
            historial_por_cuota[h.cuota_id].append(h)
    
    # También buscar cuotas que recibieron monto (fueron modificadas por un pago parcial previo)
    cuotas_con_monto_recibido = {}
    historiales_recibidos = HistorialModificacionPago.objects.filter(
        cuota_id__in=cuota_ids,
        tipo_modificacion='MR'
    ).select_related('cuota_relacionada')
    for h in historiales_recibidos:
        cuotas_con_monto_recibido[h.cuota_id] = h
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cierre {fecha.strftime('%d-%m-%Y')}"
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    modificada_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # Amarillo claro
    recibida_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # Celeste claro
    
    # Título
    ws.merge_cells('A1:S1')
    ws['A1'] = f'CIERRE DE CAJA - {fecha.strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    total_cobrado = pagos.aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
    total_efectivo = pagos.aggregate(total=Sum('monto_efectivo'))['total'] or Decimal('0.00')
    total_transferencia = pagos.aggregate(total=Sum('monto_transferencia'))['total'] or Decimal('0.00')
    ws.merge_cells('A2:S2')
    ws['A2'] = f'Total cobrado: ${total_cobrado:,.0f} (Efectivo: ${total_efectivo:,.0f} | Transferencia: ${total_transferencia:,.0f}) | Pagos: {pagos.count()} | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Leyenda de colores
    ws.merge_cells('A3:S3')
    ws['A3'] = '■ Amarillo = Pago parcial (se transfirió monto a otra cuota)  |  ■ Celeste = Cuota que recibió monto de otra cuota'
    ws['A3'].font = Font(italic=True, size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Headers - ahora con columnas de modificaciones
    headers = ['#', 'Préstamo', 'Cliente', 'Dirección', 'Teléfono', 'Cuota', 'Monto Cuota', 'Cobrado', 
               'Método Pago', 'Efectivo', 'Transferencia', 'Estado', 'Fecha Inicio', 
               '% Interés', 'Fecha Fin Préstamo', 'Cobrador',
               'Modificada', 'Monto Original', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Anchos
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 16
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 16
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 14
    ws.column_dimensions['R'].width = 16
    ws.column_dimensions['S'].width = 45
    
    # Datos
    total = Decimal('0.00')
    for i, pago in enumerate(pagos, 1):
        row = i + 5
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=f'#{pago.prestamo.pk}').border = border
        ws.cell(row=row, column=3, value=pago.prestamo.cliente.nombre_completo).border = border
        ws.cell(row=row, column=4, value=pago.prestamo.cliente.direccion or '-').border = border
        ws.cell(row=row, column=5, value=pago.prestamo.cliente.telefono).border = border
        ws.cell(row=row, column=6, value=f'{pago.numero_cuota}/{pago.prestamo.cuotas_pactadas}').border = border
        
        monto_cell = ws.cell(row=row, column=7, value=float(pago.monto_cuota))
        monto_cell.number_format = '#,##0'
        monto_cell.border = border
        
        cobrado_cell = ws.cell(row=row, column=8, value=float(pago.monto_pagado))
        cobrado_cell.number_format = '#,##0'
        cobrado_cell.border = border
        cobrado_cell.font = Font(bold=True, color='198754')
        
        ws.cell(row=row, column=9, value=pago.get_metodo_pago_display()).border = border
        
        ef_cell = ws.cell(row=row, column=10, value=float(pago.monto_efectivo or 0))
        ef_cell.number_format = '#,##0'
        ef_cell.border = border
        
        tr_cell = ws.cell(row=row, column=11, value=float(pago.monto_transferencia or 0))
        tr_cell.number_format = '#,##0'
        tr_cell.border = border
        
        ws.cell(row=row, column=12, value=pago.get_estado_display()).border = border
        ws.cell(row=row, column=13, value=pago.prestamo.fecha_inicio.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=14, value=f'{pago.prestamo.tasa_interes_porcentaje}%').border = border
        ws.cell(row=row, column=15, value=pago.prestamo.fecha_finalizacion.strftime('%d/%m/%Y') if pago.prestamo.fecha_finalizacion else '-').border = border
        ws.cell(row=row, column=16, value=pago.cobrado_por.get_full_name() or pago.cobrado_por.username if pago.cobrado_por else '-').border = border
        
        # --- Columnas de Modificaciones ---
        historial = historial_por_cuota.get(pago.id, [])
        recibido = cuotas_con_monto_recibido.get(pago.id)
        
        fue_modificada = False
        monto_original = ''
        observaciones_parts = []
        row_fill = None
        
        for h in historial:
            if h.tipo_modificacion == 'PP':
                fue_modificada = True
                row_fill = modificada_fill
                if h.monto_restante_transferido > 0:
                    observaciones_parts.append(
                        f'Pago parcial: cobrado ${h.monto_pagado:,.0f} de ${h.monto_cuota_anterior:,.0f}. '
                        f'Restante ${h.monto_restante_transferido:,.0f} transferido'
                    )
                else:
                    observaciones_parts.append(
                        f'Pago parcial: cobrado ${h.monto_pagado:,.0f} de ${h.monto_cuota_anterior:,.0f}'
                    )
            elif h.tipo_modificacion == 'TR':
                destino = f' a cuota #{h.cuota_relacionada.numero_cuota}' if h.cuota_relacionada else ''
                observaciones_parts.append(
                    f'Transferido ${h.monto_restante_transferido:,.0f}{destino}'
                )
                if h.interes_mora > 0:
                    observaciones_parts.append(f'(incluye mora: ${h.interes_mora:,.0f})')
            elif h.tipo_modificacion == 'CE':
                destino = f' (cuota #{h.cuota_relacionada.numero_cuota})' if h.cuota_relacionada else ''
                observaciones_parts.append(
                    f'Cuota especial creada por ${h.monto_restante_transferido:,.0f}{destino}'
                )
        
        if recibido:
            fue_modificada = True
            if not row_fill:
                row_fill = recibida_fill
            monto_original = float(recibido.monto_cuota_anterior)
            origen = f' de cuota #{recibido.cuota_relacionada.numero_cuota}' if recibido.cuota_relacionada else ''
            observaciones_parts.insert(0,
                f'Recibió ${recibido.monto_restante_transferido:,.0f}{origen}'
            )
            if recibido.interes_mora > 0:
                observaciones_parts.insert(1, f'(incluye mora: ${recibido.interes_mora:,.0f})')
        
        mod_cell = ws.cell(row=row, column=17, value='SÍ' if fue_modificada else '-')
        mod_cell.border = border
        mod_cell.alignment = Alignment(horizontal='center')
        if fue_modificada:
            mod_cell.font = Font(bold=True, color='856404')
        
        orig_cell = ws.cell(row=row, column=18, value=monto_original if monto_original else '-')
        if isinstance(monto_original, float):
            orig_cell.number_format = '#,##0'
        orig_cell.border = border
        
        obs_cell = ws.cell(row=row, column=19, value=' | '.join(observaciones_parts) if observaciones_parts else '-')
        obs_cell.border = border
        obs_cell.alignment = Alignment(wrap_text=True)
        
        # Aplicar color de fondo a toda la fila si fue modificada
        if row_fill:
            for col_idx in range(1, 20):
                ws.cell(row=row, column=col_idx).fill = row_fill
        
        total += pago.monto_pagado
    
    # Fila total
    total_row = pagos.count() + 6
    ws.merge_cells(f'A{total_row}:G{total_row}')
    total_label = ws.cell(row=total_row, column=1, value='TOTAL COBRADO:')
    total_label.font = Font(bold=True, size=12)
    total_cell = ws.cell(row=total_row, column=8, value=float(total))
    total_cell.font = Font(bold=True, size=12, color='198754')
    total_cell.number_format = '#,##0'
    
    # Totales efectivo y transferencia
    ef_total_cell = ws.cell(row=total_row, column=10, value=float(total_efectivo))
    ef_total_cell.font = Font(bold=True, size=11)
    ef_total_cell.number_format = '#,##0'
    tr_total_cell = ws.cell(row=total_row, column=11, value=float(total_transferencia))
    tr_total_cell.font = Font(bold=True, size=11)
    tr_total_cell.number_format = '#,##0'
    
    # Registrar auditoría
    RegistroAuditoria.registrar(
        usuario=request.user,
        tipo_accion='OT',
        tipo_modelo='SI',
        descripcion=f'Exportación de cierre de caja a Excel - Fecha: {fecha}',
        ip_address=get_client_ip(request)
    )
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=cierre_caja_{fecha.strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_clientes_excel(request):
    """Exportar lista de clientes a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, 'La exportación a Excel no está disponible. Instale openpyxl.')
        return redirect('core:cliente_list')
    
    clientes = Cliente.objects.filter(estado='AC')
    if not es_usuario_admin(request.user):
        clientes = clientes.filter(usuario=request.user)
    clientes = clientes.select_related('ruta', 'tipo_negocio')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['#', 'Nombre', 'Apellido', 'Teléfono', 'Dirección', 'Categoría', 'Ruta', 'Tipo Negocio', 'Límite Crédito']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Datos
    for i, cliente in enumerate(clientes, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=cliente.nombre).border = border
        ws.cell(row=row, column=3, value=cliente.apellido).border = border
        ws.cell(row=row, column=4, value=cliente.telefono).border = border
        ws.cell(row=row, column=5, value=cliente.direccion[:50]).border = border
        ws.cell(row=row, column=6, value=cliente.get_categoria_display()).border = border
        ws.cell(row=row, column=7, value=cliente.ruta.nombre if cliente.ruta else '-').border = border
        ws.cell(row=row, column=8, value=cliente.tipo_negocio.nombre if cliente.tipo_negocio else '-').border = border
        limite_cell = ws.cell(row=row, column=9, value=float(cliente.limite_credito))
        limite_cell.number_format = '#,##0'
        limite_cell.border = border
    
    # Ajustar anchos
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['E'].width = 30
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def exportar_prestamos_excel(request):
    """Exportar préstamos a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        messages.error(request, 'La exportación a Excel no está disponible. Instale openpyxl.')
        return redirect('core:prestamo_list')
    
    estado = request.GET.get('estado', '')
    prestamos = Prestamo.objects.select_related('cliente')
    if not es_usuario_admin(request.user):
        prestamos = prestamos.filter(cliente__usuario=request.user)
    if estado:
        prestamos = prestamos.filter(estado=estado)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['#', 'Cliente', 'Dirección', 'Monto', 'Total', 'Pagado', 'Pendiente', 'Cuotas', 'Frecuencia', 'Estado', 'Fecha Inicio', 'Fecha Finalización']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for i, p in enumerate(prestamos, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=p.cliente.nombre_completo).border = border
        ws.cell(row=row, column=3, value=p.cliente.direccion or '-').border = border
        monto_cell = ws.cell(row=row, column=4, value=float(p.monto_solicitado))
        monto_cell.number_format = '#,##0'
        monto_cell.border = border
        total_cell = ws.cell(row=row, column=5, value=float(p.monto_total_a_pagar))
        total_cell.number_format = '#,##0'
        total_cell.border = border
        pagado_cell = ws.cell(row=row, column=6, value=float(p.monto_pagado))
        pagado_cell.number_format = '#,##0'
        pagado_cell.border = border
        pend_cell = ws.cell(row=row, column=7, value=float(p.monto_pendiente))
        pend_cell.number_format = '#,##0'
        pend_cell.border = border
        ws.cell(row=row, column=8, value=f'{p.cuotas_pagadas}/{p.cuotas_pactadas}').border = border
        ws.cell(row=row, column=9, value=p.get_frecuencia_display()).border = border
        ws.cell(row=row, column=10, value=p.get_estado_display()).border = border
        ws.cell(row=row, column=11, value=p.fecha_inicio.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=12, value=p.fecha_finalizacion.strftime('%d/%m/%Y') if p.fecha_finalizacion else '-').border = border
    
    for col in range(1, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=prestamos_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


# ==================== NOTIFICACIONES ====================

class NotificacionListView(LoginRequiredMixin, ListView):
    """Vista de notificaciones del usuario"""
    model = Notificacion
    template_name = 'core/notificacion_list.html'
    context_object_name = 'notificaciones'
    paginate_by = 20
    
    def get_queryset(self):
        qs = Notificacion.objects.filter(
            Q(usuario=self.request.user) | Q(usuario__isnull=True)
        )
        
        # Filtros
        solo_no_leidas = self.request.GET.get('no_leidas', '')
        tipo = self.request.GET.get('tipo', '')
        
        if solo_no_leidas:
            qs = qs.filter(leida=False)
        if tipo:
            qs = qs.filter(tipo=tipo)
        
        return qs.order_by('-fecha_creacion')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_notificacion'] = Notificacion.TipoNotificacion.choices
        context['no_leidas_count'] = Notificacion.objects.filter(
            Q(usuario=self.request.user) | Q(usuario__isnull=True),
            leida=False
        ).count()
        return context


@login_required
def marcar_notificacion_leida(request, pk):
    """Marcar notificación como leída via AJAX"""
    notificacion = get_object_or_404(Notificacion, pk=pk)
    notificacion.marcar_como_leida()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('core:notificacion_list')


@login_required
def marcar_todas_leidas(request):
    """Marcar todas las notificaciones como leídas"""
    Notificacion.objects.filter(
        Q(usuario=request.user) | Q(usuario__isnull=True),
        leida=False
    ).update(leida=True, fecha_lectura=timezone.now())
    
    messages.success(request, 'Todas las notificaciones marcadas como leídas.')
    return redirect('core:notificacion_list')


@login_required
def obtener_notificaciones(request):
    """API para obtener notificaciones no leídas (para actualización en tiempo real)"""
    notificaciones = Notificacion.objects.filter(
        Q(usuario=request.user) | Q(usuario__isnull=True),
        leida=False
    ).order_by('-fecha_creacion')[:5]
    
    data = {
        'count': notificaciones.count(),
        'notificaciones': [
            {
                'id': n.pk,
                'titulo': n.titulo,
                'mensaje': n.mensaje[:100],
                'tipo': n.tipo,
                'prioridad': n.prioridad,
                'fecha': n.fecha_creacion.strftime('%d/%m %H:%M'),
                'enlace': n.enlace
            }
            for n in notificaciones
        ]
    }
    return JsonResponse(data)


# ==================== AUDITORÍA ====================

class AuditoriaListView(LoginRequiredMixin, ListView):
    """Vista de registros de auditoría"""
    model = RegistroAuditoria
    template_name = 'core/auditoria_list.html'
    context_object_name = 'registros'
    paginate_by = 50
    
    def dispatch(self, request, *args, **kwargs):
        if not es_usuario_admin(request.user):
            messages.error(request, 'No tienes permiso para ver el historial de auditoría.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        # Filtros
        usuario = self.request.GET.get('usuario', '')
        tipo_accion = self.request.GET.get('tipo_accion', '')
        tipo_modelo = self.request.GET.get('tipo_modelo', '')
        fecha_desde = self.request.GET.get('fecha_desde', '')
        fecha_hasta = self.request.GET.get('fecha_hasta', '')
        
        if usuario:
            qs = qs.filter(usuario_id=usuario)
        if tipo_accion:
            qs = qs.filter(tipo_accion=tipo_accion)
        if tipo_modelo:
            qs = qs.filter(tipo_modelo=tipo_modelo)
        if fecha_desde:
            qs = qs.filter(fecha_hora__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_hora__date__lte=fecha_hasta)
        
        return qs.select_related('usuario')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios'] = User.objects.all()
        context['tipos_accion'] = RegistroAuditoria.TipoAccion.choices
        context['tipos_modelo'] = RegistroAuditoria.TipoModelo.choices
        return context


# ==================== RESPALDOS ====================

@login_required
def crear_respaldo(request):
    """Crear respaldo manual de la base de datos"""
    if not es_usuario_admin(request.user):
        messages.error(request, 'Solo los administradores pueden crear respaldos.')
        return redirect('core:dashboard')
    
    import shutil
    import json
    from django.conf import settings
    from django.core import serializers
    
    try:
        # Crear directorio de respaldos si no existe
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        db_engine = settings.DATABASES['default']['ENGINE']
        
        # Verificar si es PostgreSQL o SQLite
        if 'postgresql' in db_engine:
            # PostgreSQL: exportar datos a JSON
            backup_name = f'backup_{timestamp}.json'
            backup_path = os.path.join(backup_dir, backup_name)
            
            # Exportar todos los modelos a JSON
            from core.models import (
                Cliente, Prestamo, Cuota, TipoNegocio, RutaCobro,
                ConfiguracionCredito, ConfiguracionPlanilla, PerfilUsuario,
                RegistroAuditoria, Notificacion
            )
            from django.contrib.auth.models import User
            
            all_data = {}
            models_to_export = [
                ('users', User),
                ('perfiles', PerfilUsuario),
                ('tipos_negocio', TipoNegocio),
                ('rutas_cobro', RutaCobro),
                ('config_credito', ConfiguracionCredito),
                ('config_planilla', ConfiguracionPlanilla),
                ('clientes', Cliente),
                ('prestamos', Prestamo),
                ('cuotas', Cuota),
            ]
            
            for name, model in models_to_export:
                all_data[name] = json.loads(serializers.serialize('json', model.objects.all()))
            
            with open(backup_path, 'w', encoding='utf-8') as f:
                json.dump(all_data, f, ensure_ascii=False, indent=2, default=str)
                
        else:
            # SQLite: copiar archivo
            backup_name = f'backup_{timestamp}.sqlite3'
            backup_path = os.path.join(backup_dir, backup_name)
            db_path = settings.DATABASES['default']['NAME']
            shutil.copy2(db_path, backup_path)
        
        # Registrar auditoría
        RegistroAuditoria.registrar(
            usuario=request.user,
            tipo_accion='RS',
            tipo_modelo='SI',
            descripcion=f'Respaldo manual creado: {backup_name}',
            ip_address=get_client_ip(request)
        )
        
        # Limpiar respaldos antiguos
        config = ConfiguracionRespaldo.objects.first()
        if config:
            config.ultimo_respaldo = timezone.now()
            config.save()
            
            # Mantener solo los últimos N respaldos
            backups = sorted(
                [f for f in os.listdir(backup_dir) if f.startswith('backup_')],
                reverse=True
            )
            for old_backup in backups[config.mantener_ultimos:]:
                os.remove(os.path.join(backup_dir, old_backup))
        
        messages.success(request, f'Respaldo creado exitosamente: {backup_name}')
    except Exception as e:
        messages.error(request, f'Error al crear respaldo: {str(e)}')
    
    return redirect('core:reporte_general')


@login_required
def descargar_respaldo(request, nombre):
    """Descargar un respaldo específico"""
    if not es_usuario_admin(request.user):
        messages.error(request, 'Solo los administradores pueden descargar respaldos.')
        return redirect('core:dashboard')
    
    from django.conf import settings
    
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    backup_path = os.path.join(backup_dir, nombre)
    
    if os.path.exists(backup_path) and nombre.startswith('backup_'):
        with open(backup_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename={nombre}'
            return response
    
    messages.error(request, 'El archivo de respaldo no existe.')
    return redirect('core:reporte_general')


class RespaldoListView(LoginRequiredMixin, TemplateView):
    """Vista de listado de respaldos disponibles"""
    template_name = 'core/respaldo_list.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not es_usuario_admin(request.user):
            messages.error(request, 'Solo los administradores pueden ver los respaldos.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        from django.conf import settings
        
        context = super().get_context_data(**kwargs)
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        
        backups = []
        if os.path.exists(backup_dir):
            for f in sorted(os.listdir(backup_dir), reverse=True):
                if f.startswith('backup_'):
                    path = os.path.join(backup_dir, f)
                    size = os.path.getsize(path)
                    backups.append({
                        'nombre': f,
                        'tamano': f'{size / 1024 / 1024:.2f} MB',
                        'fecha': datetime.fromtimestamp(os.path.getctime(path))
                    })
        
        context['backups'] = backups
        context['config'] = ConfiguracionRespaldo.objects.first()
        return context


# ==================== UTILIDADES ====================

def get_client_ip(request):
    """Obtener la IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# ==================== GENERAR NOTIFICACIONES AUTOMÁTICAS ====================

@login_required
def generar_notificaciones(request):
    """Generar notificaciones de cuotas vencidas y por vencer"""
    if not es_usuario_admin(request.user):
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)
    
    Notificacion.notificar_cuotas_vencidas()
    Notificacion.notificar_cuotas_por_vencer()
    
    return JsonResponse({'success': True, 'message': 'Notificaciones generadas'})
